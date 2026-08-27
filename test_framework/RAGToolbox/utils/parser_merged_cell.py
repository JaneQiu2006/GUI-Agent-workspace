import openpyxl

# 定义函数来解析合并单元格
def parser_merged_cell(sheet: openpyxl.worksheet.worksheet.Worksheet, row, col):
    """
    检查是否为合并单元格并获取对应行列单元格的值。
    如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
    :param sheet: 当前工作表对象
    :param row: 需要获取的单元格所在行
    :param col: 需要获取的单元格所在列
    :return: 单元格的值
    """
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 获取合并区域左上角的单元格作为该单元格的值返回
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell

if __name__ == "__main__":
    # 读取Excel文件
    file_path = './0mvp/车控操作树.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet_ = wb.active

    for row_index in range(2, sheet_.max_row + 1):
        row_content = []
        for col_index in range(2, sheet_.max_column + 1):
            cell_ = parser_merged_cell(sheet_, row_index, col_index)
            if cell_ and cell_.value and cell_.value != '/':
                row_content.append(cell_.value.replace("\n", ""))
        if row_content:
            print("在中控屏进入设置 > "+" > ".join(row_content))